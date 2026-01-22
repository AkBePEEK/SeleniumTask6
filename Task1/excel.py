from openpyxl import load_workbook

def read_excel_data(file_path, sheet_name=None):
    """
    Read test data from Excel file using openpyxl.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet (default: first sheet)
    
    Returns:
        List of dictionaries with test data
    """
    workbook = load_workbook(file_path)
    
    if sheet_name is None:
        worksheet = workbook.active
    else:
        worksheet = workbook[sheet_name]
    
    # Get headers from first row
    headers = [cell.value for cell in worksheet[1]]
    
    # Read data rows
    test_data = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # Skip empty rows
            row_dict = dict(zip(headers, row))
            test_data.append(row_dict)
    
    workbook.close()
    return test_data


if __name__ == "__main__":
    # Example usage
    data = read_excel_data("test_data.xlsx")
    for test_case in data:
        print(test_case)